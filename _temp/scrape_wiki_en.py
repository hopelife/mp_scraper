# [파이썬 크롤러 만들기](https://jione-e.tistory.com/28)
# [Source Code for Package lxml.html](https://lxml.de/api/lxml.html-pysrc.html)

import os, sys
import json
import time
import re
import openpyxl
# from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

import requests
import lxml.html as etree
import lxml.etree as et

sys.path.append('../_public') ## Note: 현재 디렉토리 기준 상대 경로 설정
from scrap_requests import (_root_tree, _fn)
from utils_basic import (_read_file, _create_folder, _write_file, _fn)

# url = 'https://jmoney77.tistory.com/notice/682'
# requests.get(url).source

def _normalize_spaces(s):
    """연결된 공백을 하나의 공백으로 변경
    """
    return re.sub(r'\s+', ' ', s).strip()


def _remove_punc(s):
    """ 특수문자 제거

    Args:
        s (str): 입력 문자열
    """
    punc = '[!"#$%&\'()*+,-./:;<=>?[\]^_`{|}~“”·]'
    return re.sub(punc, '', s)


def _save_file_by_url(path, url):
    """해당 url의 파일(이미지 포함)을 path에 저장
    """
    _create_folder('/'.join(path.split('/')[:-1]))
    r = requests.get(url)

    for i in range(100):  ## TODO: 이미지 request가 처리할 때까지 반복, 작동 확인!!
        if r.status_code == 200:
            # print(f"success: {i}번째 시도")
            with open(path, 'wb') as f:
                f.write(r.content)
            break
        time.sleep(0.05)


def _extract_detail_pages(url, xpath):
    """상세 페이지 title, url 추출
    """
    page = requests.get(url)
    root = etree.fromstring(page.content)
    return {
        el.xpath('.//h3/text()')[0].replace('#', '').replace('.', '').replace(' (', '('): el.xpath('./a')[0].get('href')
        for el in root.xpath(xpath)
    }


def scrape_list_pages(pages, xpath):
    """스크랩할 page 리스트
    TODO: response = session.get(url)  # Session / yield 사용
    """
    _pages = {}
    if type(pages) == list:
        for url in pages:
            _pages = dict(_pages, **_extract_detail_pages(url, xpath))
    elif type(pages) == dict:  # dict
        for sub, url in pages.items():
            _pages[sub] = {}
            _pages[sub] = dict(_pages[sub], **_extract_detail_pages(url, xpath))

    return _pages


def scrape_detail_page(title, url, base_url, base_path):
    """스크랩할 page 리스트
    url: 스크랩할 페이지
    base_url: 스크랩할 site url
    base_path: 저장할 폴더
    TODO: response = session.get(url)  # Session / yield 사용
    """
    root = etree.fromstring(requests.get(url).content)

    ## HTML wraper
    html = et.Element("html")
    head = et.Element("head")
    body = et.Element("body")

    html.insert(0, head)
    html.insert(1, body)

    ## meta(encoding)
    meta = et.Element("meta")
    meta.set('charset', 'utf-8')
    head.insert(0, meta)

    ## style
    styles = root.xpath(".//style[@type='text/css']")
    for style in styles:
        n = len(head.xpath('.//style'))
        head.insert(n+1, style)  ## NOTE: meta, 앞의 style 뒤에 추가

    ## title
    _title = root.xpath('.//*[@id="firstHeading"]')[0]
    
    root = root.xpath('//*[@id="bodyContent"]')[0]  # 확인 필요

    figs = root.xpath('.//img')

    root.make_links_absolute(base_url)  # TODO: url들을 절대 주소로 변환

    for fig in figs:
        try:    #  KeyError: 'data-filename'
            url = fig.attrib['src']
            path = f'images/' + url.split('/')[-1]

            if '?' in path:
                path = f'images/checkLoggedIn.png'

            print(f'{base_path}/{path}', url)
            _save_file_by_url(f'{base_path}/{path}', url)  # NOTE: image 파일 저장
            time.sleep(0.1)
            ## NOTE: figure element -> img element
            # fig.set('src', path)

            img = et.Element("img")
            img.set('src', path)
            img.set('alt', fig.get('alt'))
            img.set('width', fig.get('width'))
            img.set('height', fig.get('height'))
            fig.getparent().replace(fig, img)  ## NOTE: element 치환(figure -> img) 
        except:
            print(f"fig: {etree.tostring(fig)}")
            pass

    body.insert(0, _title)
    body.insert(1, root)

    open(f'{base_path }/{title}.html', 'w', encoding="utf-8").write(_normalize_spaces(etree.tostring(html, encoding=str))) ## NOTE: html 저장, endcoding=str, spaces 정리  ## html 저장


if __name__ == '__main__':

    base_url = 'https://en.wikipedia.org'
    base_path = 'wikipedia_org'
    prefix = f'{base_url}/wiki'
    titles = [
        'Candlestick_pattern',
    ]

    for title in titles:
        url = f'{prefix}/{title}'
        scrape_detail_page(title, url, base_url, base_path)
        time.sleep(0.1)


    # ## NOTE: after scraping
    # path = 'wikipedia_org/Candlestick_pattern.html'
    # root = etree.fromstring(_read_file(path))

    # names = []
    # meanings = []
    # images = []

    # for tr in root.xpath('.//*[@id="mw-content-text"]/div[1]//tr'):
    #     tds = tr.xpath('./td')

    #     if len(tds) != 4:  # NOTE: td가 4개가 아닌 행이면
    #         continue
    #     # elif  tds[1].text == None:  # NOTE: 비어있는 행이면
    #     #     continue
    #     print(f"tds len: {len(tds)}")
    #     for i, td in enumerate(tds):
    #         try:
    #             # image = images.append(td.xpath('.//img')[0]
    #             if i%2 == 0:  # NOTE: image 포함 열
    #                 images.append(td.xpath('.//img')[0].attrib['src'])
    #             else:
    #                 names.append(td.xpath('.//b')[0].text_content())
    #                 meanings.append(td.text_content())
    #         except:
    #             print("의미없는 행")

    # zips = zip(images, names, meanings)
    # candles = []
    # for image, name, meaning in zips:
    #     candles.append({'image': image, 'name': name.replace(' Candle', '').replace(' ', '_'), 'meaning': meaning})

    # # print(candles)

    # wb = openpyxl.Workbook()
    # ws = wb.worksheets[0]
    # ws.title = 'wikipedia'

    # ## NOTE: cell header 입력
    # for i, col in enumerate(candles[0].keys()):
    #     # NOTE: name
    #     ws.cell(row=1, column=i+1, value=col)
    
    # ws.cell(row=1, column=4, value="condition")

    # ## NOTE: cell contents 입력
    # for i, candle in enumerate(candles):
    #     # col = openpyxl.utils.get_column_letter(i + 2)
    #     # NOTE: image
    #     path = f"wikipedia_org/{candle['image']}"
    #     img = openpyxl.drawing.image.Image(path)
    #     img.anchor = f'A{i+2}'
    #     ws.add_image(img)

    #     # NOTE: name
    #     ws.cell(row=i+2, column=2, value=candle['name'])

    #     # NOTE: meaning
    #     ws.cell(row=i+2, column=3, value=candle['meaning'])
    #     # ws[f'C{i+2}'].alignment.wrap_text = True

    # ## NOTE: cell style 설정

    # ## NOTE: 사용자 설정 style 050A30
    # style_header = openpyxl.styles.NamedStyle(name="style_header")
    # style_header.font = openpyxl.styles.Font(color="DAEEF3", bold=True, size=12)
    # style_header.fill = openpyxl.styles.PatternFill("solid", fgColor="050A30")
    # # style_header.fill = openpyxl.styles.PatternFill(start_color="050A30", fill_type = "solid")
    # style_header.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    # bd = openpyxl.styles.Side(style='thin', color="DAEEF3")
    # style_header.border = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
    # wb.add_named_style(style_header)

    # style_contents = openpyxl.styles.NamedStyle(name="style_contents")
    # style_contents.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
    # bd = openpyxl.styles.Side(style='thin', color="050A30")
    # style_contents.border = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
    # style_contents.font = openpyxl.styles.Font(color="050A30", size=10)
    # # style_contents.font = openpyxl.styles.Font(color="ff00ff", size=10)
    # style_contents.fill = openpyxl.styles.PatternFill("solid", fgColor="DAEEF3")
    # # style_contents.fill = openpyxl.styles.PatternFill("solid", fgColor="ff0000")
    # wb.add_named_style(style_contents)


    # ws.column_dimensions['A'].width = int(13)  # 1 unit = 7 pixels, 0.7109375
    # ws.column_dimensions['B'].width = int(25)  # 1 unit = 7 pixels, 0.7109375
    # ws.column_dimensions['C'].width = int(80)  # 1 unit = 7 pixels, 0.7109375
    # ws.column_dimensions['D'].width = int(40)  # 1 unit = 7 pixels, 0.7109375

    # for i in range(3):
    #     ws[f"{openpyxl.utils.get_column_letter(i+1)}1"].style = 'style_header'
    
    # ws["D1"].style = 'style_header'  ## NOTE: condition

    # for i in range(len(candles)):   ## NOTE: contents
    #     ws.row_dimensions[i+2].height = 100*0.75  # 12.5  # Corresponding to font size 12
    #     ws[f'A{i+2}'].style = 'style_contents'
    #     ws[f'A{i+2}'].fill = openpyxl.styles.PatternFill("solid", fgColor="FDE9D9")
    #     # bd = openpyxl.styles.Side(style='thin', color="DAEEF3")
    #     # ws[f'A{i+2}'].border = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
    #     ws[f'B{i+2}'].style = 'style_contents'
    #     ws[f'C{i+2}'].style = 'style_contents'
    #     ws[f'D{i+2}'].style = 'style_contents'
    #     # ws[f'B{i+2}'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
    #     # ws[f'C{i+2}'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ## NOTE: 저장
    # wb.save('out.xlsx')




    # width = 15 # in cm, approx 5 rows (Actually it should say columns here...)
    # height = 7.5 # in cm, approx 14 rows




    # B1 셀에 값 입력 (추천하는 방법)
    # ws.cell(row=1, column=2, value=70)

    # openpyxl.utils.get_column_letter(27) => 'AA'
